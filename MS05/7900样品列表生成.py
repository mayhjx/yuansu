# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk
import tkinter.messagebox as msg
from openpyxl import Workbook
import re
import datetime
import os, sys
import logging

def CreatePatch(start, end, table):

    Table1 = (1101, 1102, 1103, 1104, 1105, 1106, 1107, 1108, 1109, 1110, 
            1201, 1202, 1203, 1204, 1205, 1206, 1207, 1208, 1209, 1210, 
            1301, 1302, 1303, 1304, 1305, 1306, 1307, 1308, 1309, 1310, 
            1401, 1402, 1403, 1404, 1405, 1406, 1407, 1408, 1409, 1410, 
            1501, 1502, 1503, 1504, 1505, 1506, 1507, 1508, 1509, 1510, 
            1601, 1602, 1603, 1604, 1605, 1606, 1607, 1608, 1609, 1610)

    Table2 = (2101, 2102, 2103, 2104, 2105, 2106, 2107, 2108, 2109, 2110, 
            2201, 2202, 2203, 2204, 2205, 2206, 2207, 2208, 2209, 2210, 
            2301, 2302, 2303, 2304, 2305, 2306, 2307, 2308, 2309, 2310, 
            2401, 2402, 2403, 2404, 2405, 2406, 2407, 2408, 2409, 2410, 
            2501, 2502, 2503, 2504, 2505, 2506, 2507, 2508, 2509, 2510, 
            2601, 2602, 2603, 2604, 2605, 2606, 2607, 2608, 2609, 2610)
        
    Table4 = (4101, 4102, 4103, 4104, 4105, 4106, 4107, 4108, 4109, 4110, 
            4201, 4202, 4203, 4204, 4205, 4206, 4207, 4208, 4209, 4210, 
            4301, 4302, 4303, 4304, 4305, 4306, 4307, 4308, 4309, 4310, 
            4401, 4402, 4403, 4404, 4405, 4406, 4407, 4408, 4409, 4410, 
            4501, 4502, 4503, 4504, 4505, 4506, 4507, 4508, 4509, 4510, 
            4601, 4602, 4603, 4604, 4605, 4606, 4607, 4608, 4609, 4610)

    pattern = re.compile("((17){0,1}[A-Z]{0,3})(\d+)")

    if start == '':
        raise NameError("请输入开始实验号")
    elif end == '':
        raise NameError("请输入结束实验号")
    elif table == '':
        raise NameError("请输入样品架编号")

    try:
        start_letter = pattern.match(str(start)).group(1)
        start_number = pattern.match(str(start)).group(3)
    except AttributeError:
        raise NameError("无法识别开始实验号")

    try:
        end_letter = pattern.match(str(end)).group(1)
        end_number = pattern.match(str(end)).group(3)
    except AttributeError:
        raise NameError("无法识别结束实验号")

    total = int(end_number) - int(start_number) + 1

    if start_letter != end_letter:
        raise NameError("实验号的字母不一致")
    if end_number < start_number:
        raise NameError("实验号的数字有错误")
    if len(end_number) != len(start_number):
        raise NameError("实验号的长度不一致")
    if total > 60:
        raise NameError("超过60个样品")

    lists = []

    for i in range(0,total):
        # str.zfill(n)根据实验号长度不足时自动补0
        lists.append("%s%s" % (start_letter, (str(int(start_number)+i)).zfill(len(start_number))))

    if table == "1":
        return(zip(lists, Table1))
    elif table == "2":
        return(zip(lists, Table2))
    elif table == "4":
        return(zip(lists, Table4))
    else:
        raise NameError("样品架编号错误")

    
def Save(Patch):

    SavePath = "C:\\Users\\Public\\Desktop\\每日上机模板\\"
    # SavePath = "D:\\desktop\\"

    wb = Workbook()
    ws = wb.active

    Lists, Table = zip(*Patch)

    # print(App.Std.get())

    if App.Std.get() == 1:
        # print("StdOn")

        ws['A1'] = "Sample"
        ws['A2'] = "Sample"
        ws['A9'] = "Sample"
        ws['A10'] = "Sample"

        ws['B1'] = "SB"
        ws['B2'] = "SB"
        ws['B9'] = "SB"
        ws['B10'] = "SB"

        ws['C1'] = 3101
        ws['C2'] = 3101
        ws['C9'] = 3101
        ws['C10'] = 3101

        ws['A3'] = "CalBlk"
        ws['A4'] = "CalStd"
        ws['A5'] = "CalStd"
        ws['A6'] = "CalStd"
        ws['A7'] = "CalStd"
        ws['A8'] = "CalStd"

        for i in range(0,6):
            ws['C'+str(i+3)] = 3101 + i
            ws['B'+str(i+3)] = 'STD' + str(i)
            ws['E'+str(i+3)] = i+1
            ws['D'+str(i+3)] = 1

        ws['A11'] = "Sample"
        ws['A12'] = "Sample"
        ws['A13'] = "Sample"
        ws['B11'] = "CZ-"
        ws['B12'] = "QC1-"
        ws['B13'] = "QC2-"
        ws['D1'] = 20
        ws['D2'] = 20
        ws['D9'] = 20
        ws['D10'] = 20
        ws['D11'] = 20
        ws['D12'] = 20
        ws['D13'] = 20

        if Table[0] == 2101:
            ws['C11'] = 2613
            ws['C12'] = 2614
            ws['C13'] = 2615
        elif Table[0] == 1101:
            ws['C11'] = 1613
            ws['C12'] = 1614
            ws['C13'] = 1615
        elif Table[0] == 4101:
            ws['C11'] = 4613
            ws['C12'] = 4614
            ws['C13'] = 4615

        for i in range(0,len(Lists)):
            ws.cell(row=(i+14),column=1,value="Sample")
            ws.cell(row=(i+14),column=2,value=Lists[i])
            ws.cell(row=(i+14),column=3,value=Table[i])
            ws.cell(row=(i+14),column=4,value=20)

    else:

        # print("StdOff")

        ws['A1'] = "Sample"
        ws['A2'] = "Sample"
        ws['A3'] = "Sample"
        ws['B1'] = "CZ-"
        ws['B2'] = "QC1-"
        ws['B3'] = "QC2-"
        ws['D1'] = 20
        ws['D2'] = 20
        ws['D3'] = 20

        if Table[0] == 2101:
            ws['C1'] = 2613
            ws['C2'] = 2614
            ws['C3'] = 2615
        elif Table[0] == 1101:
            ws['C1'] = 1613
            ws['C2'] = 1614
            ws['C3'] = 1615
        elif Table[0] == 4101:
            ws['C1'] = 4613
            ws['C2'] = 4614
            ws['C3'] = 4615

        for i in range(0,len(Lists)):
            ws.cell(row=(i+4),column=1,value="Sample")
            ws.cell(row=(i+4),column=2,value=Lists[i])
            ws.cell(row=(i+4),column=3,value=Table[i])
            ws.cell(row=(i+4),column=4,value=20)

    now = datetime.datetime.strftime(datetime.datetime.now(),"%Y%m%d %H%M")
    
    if not os.path.isdir(SavePath):
        os.mkdir(SavePath)

    wb.save(SavePath+str(Lists[0])+"-"+str(Lists[-1])+"-"+str(Table[0])[0]+" & "+now+".xlsx")


def log(info):
    
    logging.basicConfig(level=logging.DEBUG,
                        format='%(asctime)s %(message)s',
                        datefmt='%Y%m%d %H:%M:%S',
                        filename='Patch.log',
                        filemode='a')

    logging.info(info)
    logging.shutdown()


def Main():
    
    start = App.Start_Text.get()
    end = App.End_Text.get()
    table = App.Table.get()
    
    if App.Std.get() == 1:
        info = '批号：%s - %s，样品架编号：%s，全血七项标准曲线：有' % (start,end,table)
    else:
        info = '批号：%s - %s，样品架编号：%s' % (start,end,table)

    try:
        patch = CreatePatch(start, end, table)
        if msg.askyesno(title='请确认',message=info):
            Save(patch)
            log(info)
            App.Start_Text.delete(0,len(start))
            App.End_Text.delete(0,len(end))
            App.Check.deselect()
    except NameError as err:
        msg.showerror(title='警告',message=err)
    except:
        msg.showerror(title='警告',message=sys.exc_info()[1])
        log(sys.exc_info())


class App(tk.Tk):
    
    def __init__(self):
        super().__init__()

        self.title("模板列表生成")
        self.geometry("230x200")
        self.resizable(False,False)
    
        self.Start_Label = ttk.Label(self, text="开始实验号：")
        self.Start_Label.place(relx=.25,rely=.15,anchor=tk.CENTER)

        self.End_Label = ttk.Label(self, text="结束实验号：")
        self.End_Label.place(relx=.25,rely=.35,anchor=tk.CENTER)

        self.Samp_Label = ttk.Label(self, text="样品架编号：")
        self.Samp_Label.place(relx=.25,rely=.55,anchor=tk.CENTER)

        self.start = tk.StringVar()
        self.end = tk.StringVar()
        self.table = tk.StringVar()

        # 将输入的字母转换成大写
        def caps(event): 
            self.start.set(self.start.get().upper())
            self.end.set(self.end.get().upper())

        # 开始实验号
        self.Start_Text = ttk.Entry(self,width=13,textvariable=self.start)
        self.Start_Text.place(relx=.66,rely=.15,anchor=tk.CENTER)
        self.Start_Text.bind('<KeyRelease>', caps)
        self.Start_Text.focus()

        # 结束实验号
        self.End_Text = ttk.Entry(self,width=13,textvariable=self.end)
        self.End_Text.place(relx=.66,rely=.35,anchor=tk.CENTER)
        self.End_Text.bind('<KeyRelease>', caps)

        # 样品架编号
        self.Table = ttk.Combobox(self,width=5,textvariable=self.table,state='readonly')
        self.Table['values'] = (1,2,4)
        self.Table.place(relx=.58,rely=.55,anchor=tk.CENTER)
        self.Table.current(1)

        # 标准曲线
        self.Std = tk.IntVar()
        self.Check = tk.Checkbutton(self, text="全血七项标准曲线",variable=self.Std)
        self.Check.deselect()
        self.Check.place(relx=.4,rely=.7,anchor=tk.CENTER)

        self.Creat_Button = ttk.Button(self,width=9,text="生成",command=Main)
        self.Creat_Button.place(relx=.3,rely=.86,anchor=tk.CENTER)

        self.Quit_Button = ttk.Button(self,width=9,text="退出",command=self.destroy)
        self.Quit_Button.place(relx=.7,rely=.86,anchor=tk.CENTER)

        def E(event):
            Main()
        
        self.bind("<Return>",E) 

if __name__ == "__main__":
    App = App()
    App.mainloop()
